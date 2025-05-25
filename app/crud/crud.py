from http.client import HTTPException
from msilib.schema import File
from typing import Optional
from fastapi import UploadFile, Depends
from io import BytesIO
import openpyxl
from typing import Optional, Dict
import openpyxl
from fastapi import UploadFile
from sqlalchemy.orm import Session
from starlette.responses import StreamingResponse

from .. import crud
from ..database import get_db
from ..models import models
from ..models.models import *
from ..schemas import schemas
from ..schemas.schemas import *

# Faculty CRUD
def get_faculty(db: Session, faculty_id: int):
    return db.query(models.Faculty).filter(models.Faculty.id == faculty_id).first()

def get_faculties(db: Session, skip: int = 0, limit: int = 100):
    return db.query(models.Faculty).offset(skip).limit(limit).all()

def create_faculty(db: Session, faculty: schemas.FacultyCreate):
    db_faculty = models.Faculty(name=faculty.name)
    db.add(db_faculty)
    db.commit()
    db.refresh(db_faculty)
    return db_faculty

def update_faculty(db: Session, faculty_id: int, faculty: schemas.FacultyUpdate):
    db_faculty = db.query(models.Faculty).filter(models.Faculty.id == faculty_id).first()
    if db_faculty:
        db_faculty.name = faculty.name
        db.commit()
        db.refresh(db_faculty)
    return db_faculty

def delete_faculty(db: Session, faculty_id: int):
    db_faculty = db.query(models.Faculty).filter(models.Faculty.id == faculty_id).first()
    if db_faculty:
        db.delete(db_faculty)
        db.commit()
        return True
    return False


# Program CRUD
def get_program(db: Session, program_id: int):
    return db.query(models.Program).filter(models.Program.id == program_id).first()


def get_programs(db: Session, skip: int = 0, limit: int = 100):
    return db.query(models.Program).offset(skip).limit(limit).all()


def create_program(db: Session, program: schemas.ProgramCreate):
    db_program = models.Program(**program.model_dump())
    db.add(db_program)
    db.commit()
    db.refresh(db_program)
    return db_program


def update_program(db: Session, program_id: int, program: schemas.ProgramUpdate):
    db_program = db.query(models.Program).filter(models.Program.id == program_id).first()
    if db_program:
        for key, value in program.model_dump().items():
            setattr(db_program, key, value)
        db.commit()
        db.refresh(db_program)
    return db_program


def delete_program(db: Session, program_id: int):
    db_program = db.query(models.Program).filter(models.Program.id == program_id).first()
    if db_program:
        db.delete(db_program)
        db.commit()
        return True
    return False


# Group CRUD
def get_group(db: Session, group_id: int):
    return db.query(models.Group).filter(models.Group.id == group_id).first()


def get_groups(db: Session, skip: int = 0, limit: int = 100):
    return db.query(models.Group).offset(skip).limit(limit).all()


def create_group(db: Session, group: schemas.GroupCreate):
    db_group = models.Group(**group.model_dump())
    db.add(db_group)
    db.commit()
    db.refresh(db_group)
    return db_group


def update_group(db: Session, group_id: int, group: schemas.GroupUpdate):
    db_group = db.query(models.Group).filter(models.Group.id == group_id).first()
    if db_group:
        for key, value in group.model_dump().items():
            setattr(db_group, key, value)
        db.commit()
        db.refresh(db_group)
    return db_group


def delete_group(db: Session, group_id: int):
    db_group = db.query(models.Group).filter(models.Group.id == group_id).first()
    if db_group:
        db.delete(db_group)
        db.commit()
        return True
    return False


# Student CRUD
def get_student(db: Session, student_id: int):
    return db.query(models.Student).filter(models.Student.id == student_id).first()


def get_students(db: Session, skip: int = 0, limit: int = 100):
    return db.query(models.Student).offset(skip).limit(limit).all()


def create_student(db: Session, student: schemas.StudentCreate):
    db_student = models.Student(**student.model_dump())
    db.add(db_student)
    db.commit()
    db.refresh(db_student)
    return db_student


def update_student(db: Session, student_id: int, student: schemas.StudentUpdate):
    db_student = db.query(models.Student).filter(models.Student.id == student_id).first()
    if db_student:
        for key, value in student.model_dump().items():
            setattr(db_student, key, value)
        db.commit()
        db.refresh(db_student)
    return db_student


def delete_student(db: Session, student_id: int):
    db_student = db.query(models.Student).filter(models.Student.id == student_id).first()
    if db_student:
        db.delete(db_student)
        db.commit()
        return True
    return False


# Добавим в Faculty CRUD
def get_faculty_programs(db: Session, faculty_id: int):
    return db.query(models.Program).filter(models.Program.faculty_id == faculty_id).all()


# Добавим в Program CRUD
def get_program_groups(db: Session, program_id: int):
    return db.query(models.Group).filter(models.Group.program_id == program_id).all()


# Добавим в Group CRUD
def get_group_students(db: Session, group_id: int,
                       budget_contract: Optional[str] = None,
                       status: Optional[str] = None,
                       gender: Optional[str] = None):
    query = db.query(models.Student).filter(models.Student.group_id == group_id)

    if budget_contract:
        query = query.filter(models.Student.budget_contract == budget_contract)
    if status:
        query = query.filter(models.Student.status == status)
    if gender:
        query = query.filter(models.Student.gender == gender)

    return query.all()


def export_entities_to_excel(db: Session, entity_type: str, include_ids: bool = True):
    workbook = openpyxl.Workbook()
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])

    if entity_type == 'faculty':
        data = get_faculties(db)
        sheet = workbook.create_sheet('Faculties')
        headers = ['ID', 'Name'] if include_ids else ['Name']
        sheet.append(headers)
        for item in data:
            row = [item.id, item.name] if include_ids else [item.name]
            sheet.append(row)

    elif entity_type == 'program':
        data = get_programs(db)
        sheet = workbook.create_sheet('Programs')
        headers = ['ID', 'Name', 'Faculty ID', 'Education Level'] if include_ids else ['Name', 'Faculty ID',
                                                                                       'Education Level']
        sheet.append(headers)
        for item in data:
            row = [item.id, item.name, item.faculty_id, item.education_level] if include_ids else [item.name,
                                                                                                   item.faculty_id,
                                                                                                   item.education_level]
            sheet.append(row)

    elif entity_type == 'group':
        data = get_groups(db)
        sheet = workbook.create_sheet('Groups')
        headers = ['ID', 'Name', 'Program ID', 'Study Form', 'Education Level'] if include_ids else ['Name',
                                                                                                     'Program ID',
                                                                                                     'Study Form',
                                                                                                     'Education Level']
        sheet.append(headers)
        for item in data:
            row = [item.id, item.name, item.program_id, item.study_form, item.education_level] if include_ids else [
                item.name, item.program_id, item.study_form, item.education_level]
            sheet.append(row)

    elif entity_type == 'student':
        data = get_students(db)
        sheet = workbook.create_sheet('Students')
        headers = [
            'ID', 'Last Name', 'First Name', 'Middle Name', 'Birth Date',
            'Email', 'Phone Number', 'Address', 'Gender', 'Budget/Contract',
            'Status', 'Group ID', 'Education Level'
        ] if include_ids else [
            'Last Name', 'First Name', 'Middle Name', 'Birth Date',
            'Email', 'Phone Number', 'Address', 'Gender', 'Budget/Contract',
            'Status', 'Group ID', 'Education Level'
        ]
        sheet.append(headers)
        for item in data:
            row = [
                item.id, item.last_name, item.first_name, item.middle_name,
                item.birth_date, item.email, item.phone_number, item.address,
                item.gender, item.budget_contract, item.status, item.group_id,
                item.education_level
            ] if include_ids else [
                item.last_name, item.first_name, item.middle_name,
                item.birth_date, item.email, item.phone_number, item.address,
                item.gender, item.budget_contract, item.status, item.group_id,
                item.education_level
            ]
            sheet.append(row)

    else:
        raise ValueError(f"Unknown entity type: {entity_type}")

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def import_entities_from_excel(db: Session, file: UploadFile):
    result = {
        "message": "Import completed",
        "total_records": 0,
        "imported_records": 0,
        "skipped_records": 0,
        "errors": []
    }

    try:
        workbook = openpyxl.load_workbook(filename=BytesIO(file.file.read()))

        # Process Faculties
        if 'Faculties' in workbook.sheetnames:
            sheet = workbook['Faculties']
            headers = [cell.value for cell in sheet[1]]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue

                result["total_records"] += 1
                try:
                    faculty_data = {"name": row[headers.index('Name')]}
                    create_faculty(db=db, faculty=schemas.FacultyCreate(**faculty_data))
                    result["imported_records"] += 1
                except Exception as e:
                    result["skipped_records"] += 1
                    result["errors"].append(f"Faculty row {row}: {str(e)}")

        # Process Programs
        if 'Programs' in workbook.sheetnames:
            sheet = workbook['Programs']
            headers = [cell.value for cell in sheet[1]]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue

                result["total_records"] += 1
                try:
                    program_data = {
                        "name": row[headers.index('Name')],
                        "faculty_id": row[headers.index('Faculty ID')],
                        "education_level": row[headers.index('Education Level')]
                    }
                    create_program(db=db, program=schemas.ProgramCreate(**program_data))
                    result["imported_records"] += 1
                except Exception as e:
                    result["skipped_records"] += 1
                    result["errors"].append(f"Program row {row}: {str(e)}")

        # Process Groups
        if 'Groups' in workbook.sheetnames:
            sheet = workbook['Groups']
            headers = [cell.value for cell in sheet[1]]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue

                result["total_records"] += 1
                try:
                    group_data = {
                        "name": row[headers.index('Name')],
                        "program_id": row[headers.index('Program ID')],
                        "study_form": row[headers.index('Study Form')],
                        "education_level": row[headers.index('Education Level')]
                    }
                    create_group(db=db, group=schemas.GroupCreate(**group_data))
                    result["imported_records"] += 1
                except Exception as e:
                    result["skipped_records"] += 1
                    result["errors"].append(f"Group row {row}: {str(e)}")

        # Process Students
        if 'Students' in workbook.sheetnames:
            sheet = workbook['Students']
            headers = [cell.value for cell in sheet[1]]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue

                result["total_records"] += 1
                try:
                    student_data = {
                        "last_name": row[headers.index('Last Name')],
                        "first_name": row[headers.index('First Name')],
                        "middle_name": row[headers.index('Middle Name')],
                        "birth_date": row[headers.index('Birth Date')],
                        "email": row[headers.index('Email')],
                        "phone_number": row[headers.index('Phone Number')],
                        "address": row[headers.index('Address')],
                        "gender": row[headers.index('Gender')],
                        "budget_contract": row[headers.index('Budget/Contract')],
                        "status": row[headers.index('Status')],
                        "group_id": row[headers.index('Group ID')],
                        "education_level": row[headers.index('Education Level')]
                    }
                    create_student(db=db, student=schemas.StudentCreate(**student_data))
                    result["imported_records"] += 1
                except Exception as e:
                    result["skipped_records"] += 1
                    result["errors"].append(f"Student row {row}: {str(e)}")

        return result

    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))